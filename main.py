import PySimpleGUI as sg
import datetime
import os
import pyodbc
import tkinter as tk
from tkinter import filedialog
import ctypes
import configparser as config
from openpyxl.styles import Font
from openpyxl.styles import *
import os
import openpyxl
from openpyxl import Workbook
import pandas as pd


import datetime
##* Datos generales y creación del directorio de las ventas en caso de no existir, también locación de la bd 
config = config.ConfigParser()
config.read("config.ini")
sg.theme("Reddit")

def vaciar_tabla(cursor):
    query = """DELETE FROM cartas"""
    cursor.execute(query)
    cursor.commit()
def vacio(codigo, tipo):
    if codigo == "" or tipo == "":
        return True
def existe(cursor, codigo):
    query = """SELECT * FROM cartas WHERE codigo = '{}'""".format(codigo)
    q = cursor.execute(query)
    rows = q.fetchall()
    if len(rows) == 0:
        return False
    else:
        return True
def removerCarta(cursor, codigo):
    if not existe(cursor, codigo):
        return False
    else:
        query = """DELETE FROM cartas WHERE codigo = '{}'""".format(codigo)
        cursor.execute(query)
        cursor.commit()
        return True
def connect():
    
    MDB = os.path.abspath("db_cartas.accdb")
    #Conn to DB
    DRV = "{Microsoft Access Driver (*.mdb, *.accdb)}"
    
    PWD = ""
    con = pyodbc.connect('DRIVER={};DBQ={};PWD={}'.format(DRV,MDB,PWD))
    cursor = con.cursor()
    return cursor

def getCartas(cursor):
    query = """SELECT * FROM cartas"""
    q = cursor.execute(query)
    rows = q.fetchall()
    listaCartas = []
    for i in rows:
        cartas = []
        cartas.append(i[0])
        cartas.append(i[1])
        listaCartas.append(cartas)

    return listaCartas

def añadirCarta(cursor, codigo, tipo):
    query = """INSERT INTO cartas VALUES(?,?)"""
    cursor.execute(query,(codigo, tipo))
    cursor.commit()

def exportar_Cartas(cursor):
    #lista[0] = codigo [1] = letra
    lista_de_cartas = getCartas(cursor)

    cartas_J = []
    cartas_E = []
    cartas_CM = []
    cartas_H = []
    cartas_A = []
    cartas_G = []
    cartas_F = []
    cartas_B = []
    cartas_P = []
    cartas_D = []
    cartas_K = []
    cartas_C = []

    contador = 0

    for carta in lista_de_cartas:
        if carta[1] == "j":
            cartas_J.append(carta[0])
        elif carta[1] == "e":
            cartas_E.append(carta[0])
        elif carta[1] == "c":
            cartas_C.append(carta[0])
        elif carta[1] == "cm":
            cartas_CM.append(carta[0])
        elif carta[1] == "h":
            cartas_H.append(carta[0])
        elif carta[1] == "a":
            cartas_A.append(carta[0])
        elif carta[1] == "g":
            cartas_G.append(carta[0])
        elif carta[1] == "f":
            cartas_F.append(carta[0])
        elif carta[1] == "b":
            cartas_B.append(carta[0])
        elif carta[1] == "p":
            cartas_P.append(carta[0])
        elif carta[1] == "d":
            cartas_D.append(carta[0])
        elif carta[1] == "k":
            cartas_K.append(carta[0])
        else:
            print(carta[0])
            print(carta[1])

    # 1 p
    data = [
        cartas_J,
        cartas_E,
        cartas_CM,
        cartas_H,
        cartas_A,
        cartas_G,
        cartas_F,
        cartas_B,
        cartas_P,
        cartas_D,
        cartas_K,
        cartas_C
    ]
    cartasTotales = len(cartas_C) + len(cartas_J) + len(cartas_E) + len(cartas_CM) + len(cartas_H) + len(
        cartas_A) + len(cartas_G) + len(cartas_F) + len(cartas_B) + len(cartas_P) + len(cartas_D) + len(cartas_K)
    print(len(cartas_C), "c", len(cartas_J), "j", len(cartas_E), "e", len(cartas_CM), "cm", len(cartas_H), "h",
          len(cartas_A), "a", len(cartas_G), "g", len(cartas_F), "f", len(cartas_B), "b", len(cartas_P), "p",
          len(cartas_D), "d", len(cartas_K), "k")

    nombresCartas = [
        'cartas J',
        'cartas E',
        'cartas CM',
        'cartas H',
        'cartas A',
        'cartas G',
        'cartas F',
        'cartas B',
        'cartas P',
        'cartas D',
        'cartas K',
        "cartas C"

    ]

    wb = openpyxl.Workbook()
    ws = wb.active

    extra = 0

    # DBE54E
    fuenteTituloNumero = Font(size=12, bold=True)
    borde = Border(left=Side(style='medium'), right=Side(style='medium'),
                   top=Side(style='medium'), bottom=Side(style='medium'))
    amarillo = openpyxl.styles.colors.Color(rgb='DBE54E')
    fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=amarillo)

    import configparser as config
    config = config.ConfigParser()
    config.read("config.ini")

    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['Q'].width = 20
    ws.column_dimensions['T'].width = 20
    ws.column_dimensions['W'].width = 20
    ws.column_dimensions['Z'].width = 20
    ws.column_dimensions['AC'].width = 20
    ws.column_dimensions['AF'].width = 20
    ws.column_dimensions['AI'].width = 20
    ws.column_dimensions['AL'].width = 20

    ws['E2'] = "Periodo:"
    ws['E3'] = "Usuario:"
    ws['E4'] = "Contraseña:"
    ws['E5'] = "Traductor:"
    ws['E6'] = "Total de Cartas:"

    ws['E2'].font = fuenteTituloNumero
    ws['E3'].font = fuenteTituloNumero
    ws['E4'].font = fuenteTituloNumero
    ws['E5'].font = fuenteTituloNumero
    ws['E6'].font = fuenteTituloNumero

    ws['F2'] = "ENERO - FEBRERO"
    ws['F2'].font = fuenteTituloNumero
    ws['F3'] = config["DATOS"]['usuario']
    ws['F4'] = config["DATOS"]['contrasenia']
    ws['F5'] = config["DATOS"]['traductor']
    ws['F6'] = cartasTotales

    for col, value in enumerate(data):
        cell = ws.cell(row=9, column=col + 5 + extra, value=nombresCartas[col])
        cell.font = fuenteTituloNumero
        cell.alignment = Alignment(horizontal="center")
        cell.fill = fill
        cell.border = borde
        Fill()
        print(col, value)
        for row, codigo in enumerate(value):
            enumeracion = ws.cell(row=row + 10, column=col + 4 + extra, value=row + 1)
            enumeracion.font = fuenteTituloNumero
            enumeracion.alignment = Alignment(horizontal="center")
            codigos = ws.cell(row=row + 10, column=col + 5 + extra, value=codigo)
            codigos.border = borde
            codigos.alignment = Alignment(horizontal="center")
            row += 1
        extra += 2
    # save the workbook
    wb.save('Registro de Cartas.xlsx')


cursor = connect()

cartas = getCartas(cursor)
contador = len(cartas)

def estimar_ganancia(cartas):
    ganancia_estimada = 0
    for carta in cartas:
        if carta[1] == "j":
            ganancia_estimada += 1
        elif carta[1] == "e":
            ganancia_estimada += 1
        elif carta[1] == "c":
            ganancia_estimada += 0.50
        elif carta[1] == "cm":
            ganancia_estimada += 0.50
        elif carta[1] == "h":
            ganancia_estimada += 1.10
        elif carta[1] == "a":
            ganancia_estimada += 0.50
        elif carta[1] == "g":
            ganancia_estimada += 1.10
        elif carta[1] == "f":
            ganancia_estimada += 1.10
        elif carta[1] == "b":
            ganancia_estimada += 0.50
        elif carta[1] == "p":
            ganancia_estimada += 1.10
        elif carta[1] == "d":
            ganancia_estimada += 0.50
        elif carta[1] == "k":
            ganancia_estimada += 1
    return ganancia_estimada

ganancia_estimada = estimar_ganancia(cartas)

headings = ['Código', 'Tipo Carta']
letras_de_cartas = ["j","e","c","cm","h","a","g","f","b","p","d","k"]
layout = [
        [sg.Text("Código"), sg.Input(key = "-CODIGO-"), sg.Text("Tipo"), sg.Combo(values = letras_de_cartas,key = "-TIPO-")],
        [sg.Button("Añadir", key = "-AÑADIR-"), sg.Button("Borrar Carta", key = "-BORRAR-"), sg.Button("Nuevo Registro", key = "-NUEVO-REGISTRO-"), sg.Button("Exportar", key = "-EXPORTAR-"), sg.Text("Ganancia Estimada"),sg.Text(key = "-ESTIMADO-", text = ganancia_estimada), sg.Text("Total"),sg.Text(key = "-TOTAL-",text = contador)],
        [sg.Table(values = cartas, headings=headings, 
                    max_col_width=35,
                    auto_size_columns=False,
                    justification='center',
                    key='-TABLE-',
                    enable_events=True,
                    row_height=35, expand_x = True, )]
    ]

window_lista_compra = sg.Window("Registro de Cartas", layout,modal = True, size= (1000,400))

while True:
    event, values = window_lista_compra.read()
    if event == "Exit" or event == sg.WIN_CLOSED:
        break
    elif event == "-AÑADIR-":
        if not vacio(values["-CODIGO-"], values["-TIPO-"]):
            if not existe(cursor, values["-CODIGO-"]):
                añadirCarta(cursor, values["-CODIGO-"], values ["-TIPO-"])
                cartas = getCartas(cursor)
                contador = len(cartas)
                ganancia_estimada = estimar_ganancia(cartas)
                window_lista_compra["-ESTIMADO-"].update(ganancia_estimada)
                window_lista_compra["-TOTAL-"].update(contador)
                window_lista_compra["-TABLE-"].update(cartas)
            else:
                sg.popup("Código Existente")
        else:
            sg.popup("Campos Vacíos")
    elif event == "-BORRAR-":
        if len(values['-TABLE-']) != 0:
            selected_index = values['-TABLE-'][0]
            selected_row = cartas[selected_index]
            respuesta = sg.popup_ok_cancel(f"¿Está seguro de eliminar el siguiente código?:  {selected_row[0]}",modal = True)
            if respuesta == "OK":
                removerCarta(cursor,selected_row[0])
                cartas = getCartas(cursor)
                contador = len(cartas)
                ganancia_estimada = estimar_ganancia(cartas)
                window_lista_compra["-ESTIMADO-"].update(ganancia_estimada)
                window_lista_compra["-TOTAL-"].update(contador)
                window_lista_compra["-TABLE-"].update(cartas)
    elif event == "-NUEVO-REGISTRO-":
        respuesta = sg.popup_ok_cancel(f"¿Está seguro de eliminar TODO EL REGISTRO?",modal = True)
        if respuesta == "OK":
            vaciar_tabla(cursor)
            cartas = getCartas(cursor)
            contador = len(cartas)
            ganancia_estimada = estimar_ganancia(cartas)
            window_lista_compra["-ESTIMADO-"].update(ganancia_estimada)
            window_lista_compra["-TOTAL-"].update(contador)
            window_lista_compra["-TABLE-"].update(cartas)

    elif event == "-EXPORTAR-":
        exportar_Cartas(cursor)
window_lista_compra.close()